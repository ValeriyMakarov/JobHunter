import logging
from abc import ABC, abstractmethod
from collections import Counter
from copy import deepcopy
from typing import Union, Self, Iterable

from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from src import SheetNames, CompaniesBlacklistColumnNames, \
    ApplicationColumnNames, DeniedVacanciesColumnNames, CompanyBlacklistRow, \
    DeniedVacancyRow, ApplicationDataRow
from src import TableSchemaError

log = logging.getLogger(__name__)


class BaseTable(ABC):
    class RowView:
        def __init__(self, table: Self):
            self._sheet = table._sheet
            self._table_schema = table._table_schema
            self.index = 1

        def set_index(self, index):
            self.index = index

        def __getitem__(self, column_name: Union[
                ApplicationColumnNames | DeniedVacanciesColumnNames
                | CompaniesBlacklistColumnNames
            ]
        ) -> Cell:
            return self._sheet[
                f"{self._table_schema[column_name.value]}{self.index}"
            ]

    def __init__(self, sheet: Worksheet, sheet_config: SheetNames):
        self._sheet = sheet
        self._sheet_name = sheet_config.sheet_name
        self._sheet_columns = sheet_config.columns_enum
        self._table_schema: dict[str, str] = self._build_table_schema(sheet_config)
        self._trim_table()
        self.__row = BaseTable.RowView(self)
        self._data_rows: list = []
        self._ids: set | None = None
        self._names: set | None = None
        self._read_data()

    @property
    def rows(self):
        return deepcopy(self._data_rows)

    def __contains__(self, item: str):
        return item in self._ids if self._ids else item in self._names

    def _build_table_schema(self, sheet_config: SheetNames) -> dict[str, str]:
        header_row: tuple[Cell] = self._sheet[1]
        enum_values = [item.value for item in sheet_config.columns_enum]
        schema_dict = {}
        for cell in header_row:
            if cell.value in schema_dict:
                raise TableSchemaError(
                    f"Invalid table structure. Duplicated column '{cell.value}'."
                )
            if cell.value in enum_values:
                schema_dict[cell.value] = cell.column_letter

        if not len(sheet_config.columns_enum) == len(schema_dict):
            columns_part = ", ".join(
                f"'{item.value}'"
                for item in sheet_config.columns_enum
                if item.value not in schema_dict
            )
            raise TableSchemaError(
                f"Invalid table structure. Table name = {sheet_config.sheet_name}. "
                f"Absent columns: {columns_part}.")

        return schema_dict

    def _trim_table(self):
        def row_is_empty(row: tuple):
            for col_letter in self._table_schema.values():
                if row[column_index_from_string(col_letter)].value:
                    return False
            return True

        last_column_index = max(
            column_index_from_string(col_letter)
            for col_letter in self._table_schema.values()
        )
        redundant_columns_amount = self._sheet.max_column - last_column_index
        if redundant_columns_amount > 0:
            self._sheet.delete_cols(last_column_index+1, redundant_columns_amount)

        last_row_index = self._sheet.max_row
        for i in range(self._sheet.max_row, 1, -1):
            if row_is_empty(self._sheet[i]):
                continue
            else:
                last_row_index = i
                break
        redundant_rows_amount = self._sheet.max_row - last_row_index
        if redundant_rows_amount > 0:
            self._sheet.delete_rows(last_row_index + 1, redundant_rows_amount)

    @abstractmethod
    def _read_data(self):
        ...

    @abstractmethod
    def update(
            self,
            rows: Iterable[
                ApplicationDataRow | DeniedVacancyRow | CompanyBlacklistRow
                ]
    ):
        ...

    def _get_row(self, index: int):
        self.__row.set_index(index)
        return self.__row

    def _get_duplicates(self, values: Iterable[str]) -> list[str]:
        return [k for k, v in Counter(values).items() if v > 1]


class Applications(BaseTable):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet, SheetNames.APPLICATIONS)

    @property
    def ids(self):
        return set(self._ids)

    def _read_data(self):
        applications_data = []
        applications_ids = set()
        for i in range(2, self._sheet.max_row + 1):
            row_view = self._get_row(i)

            application_row = ApplicationDataRow(
                application_date=row_view[ApplicationColumnNames.APPLICATION_DATE].value,
                vacancy_name=row_view[ApplicationColumnNames.VACANCY_NAME].value,
                company_name=row_view[ApplicationColumnNames.COMPANY_NAME].value,
                vacancy_link=row_view[ApplicationColumnNames.VACANCY_LINK].value,
                company_link=row_view[ApplicationColumnNames.COMPANY_LINK].value,
                job_description=row_view[ApplicationColumnNames.JOB_DESCRIPTION].value,
                vacancy_id=row_view[ApplicationColumnNames.VACANCY_ID].value,
                status=row_view[ApplicationColumnNames.STATUS].value,
                finish_date=row_view[ApplicationColumnNames.FINISH_DATE].value,
                cv_name=row_view[ApplicationColumnNames.CV_NAME].value,
            )
            applications_data.append(application_row)
            applications_ids.add(application_row.vacancy_id)

        duplicates = self._get_duplicates(row.vacancy_id for row in applications_data)
        vacancies_ids_part = ', '.join(applications_ids)
        duplicates_part = ', '.join(duplicates)

        log.debug(
            f"Read {len(applications_data)} applied vacancies: {vacancies_ids_part}")
        if duplicates:
            log.warning(f"Applied vacancies list has duplicates: {duplicates_part}")

        self._data_rows = applications_data
        self._ids = applications_ids

    def update(self, rows: Iterable[ApplicationDataRow]):
        for row in rows:
            if row.vacancy_id not in self._ids:
                table_row = {
                    ApplicationColumnNames.APPLICATION_DATE.value: row.application_date,
                    ApplicationColumnNames.VACANCY_NAME.value: row.vacancy_name,
                    ApplicationColumnNames.COMPANY_NAME.value: row.company_name,
                    ApplicationColumnNames.VACANCY_LINK.value: row.vacancy_link,
                    ApplicationColumnNames.COMPANY_LINK.value: row.company_link,
                    ApplicationColumnNames.JOB_DESCRIPTION.value: row.job_description,
                    ApplicationColumnNames.VACANCY_ID.value: row.vacancy_id,
                    ApplicationColumnNames.STATUS.value: row.status,
                    ApplicationColumnNames.FINISH_DATE.value: row.finish_date,
                    ApplicationColumnNames.CV_NAME.value: row.cv_name
                }
                self._sheet.append(table_row)
                self._data_rows.append(row)
                self._ids.add(row.vacancy_id)
            else:
                log.warning(f"Application data duplicate ignored: id={row.vacancy_id}")


class DeniedVacancies(BaseTable):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet, SheetNames.DENIED_VACANCIES)

    @property
    def ids(self):
        return set(self._ids)

    def _read_data(self):
        denied_vacancies_data = []
        denied_vacancies_ids = set()
        for i in range(2, self._sheet.max_row + 1):
            row_view = self._get_row(i)

            denied_vacancies_row = DeniedVacancyRow(
                vacancy_id=row_view[DeniedVacanciesColumnNames.VACANCY_ID].value,
                vacancy_name=row_view[DeniedVacanciesColumnNames.VACANCY_NAME].value,
                vacancy_link=row_view[DeniedVacanciesColumnNames.VACANCY_LINK].value,
                company_name=row_view[DeniedVacanciesColumnNames.COMPANY_NAME].value,
                company_link=row_view[DeniedVacanciesColumnNames.COMPANY_LINK].value
            )
            denied_vacancies_data.append(denied_vacancies_row)
            denied_vacancies_ids.add(denied_vacancies_row.vacancy_id)

        duplicates = self._get_duplicates(row.vacancy_id for row in denied_vacancies_data)
        vacancy_ids_part = ', '.join(denied_vacancies_ids)
        duplicated_ids_part = ', '.join(duplicates)

        log.debug(
            f"Read {len(denied_vacancies_data)} denied vacancies: {vacancy_ids_part}")
        if duplicates:
            log.warning(f"Denied vacancies list has duplicates: {duplicated_ids_part}")

        self._data_rows = denied_vacancies_data
        self._ids = denied_vacancies_ids

    def update(self, rows: Iterable[DeniedVacancyRow]):
        for row in rows:
            if row.vacancy_id not in self._ids:
                table_row = {
                    DeniedVacanciesColumnNames.VACANCY_ID.value: row.vacancy_id,
                    DeniedVacanciesColumnNames.VACANCY_NAME.value: row.vacancy_name,
                    DeniedVacanciesColumnNames.VACANCY_LINK.value: row.vacancy_link,
                    DeniedVacanciesColumnNames.COMPANY_NAME.value: row.company_name,
                    DeniedVacanciesColumnNames.COMPANY_LINK.value: row.company_link,
                }
                self._sheet.append(table_row)
                self._data_rows.append(row)
                self._ids.add(row.vacancy_id)
            else:
                log.warning(f"Denied vacancy data duplicate ignored: id={row.vacancy_id}")


class CompaniesBlacklist(BaseTable):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet, SheetNames.COMPANIES_BLACKLIST)

    @property
    def names(self):
        return set(self._names)

    def _read_data(self):
        blacklist_data = []
        blacklist_names = set()
        for i in range(2, self._sheet.max_row + 1):
            row_view = self._get_row(i)

            blacklist_row = CompanyBlacklistRow(
                company_name=row_view[CompaniesBlacklistColumnNames.COMPANY_NAME].value,
                company_link=row_view[CompaniesBlacklistColumnNames.COMPANY_LINK].value,
                reason=row_view[CompaniesBlacklistColumnNames.REASON].value
            )
            blacklist_data.append(blacklist_row)
            blacklist_names.add(blacklist_row.company_name)

        duplicates = self._get_duplicates(row.company_name for row in blacklist_data)
        company_names_part = ', '.join(blacklist_names)
        duplicated_names_part = ', '.join(duplicates)

        log.debug(f"Read {len(blacklist_data)} blacklisted companies: {company_names_part}")
        if duplicates:
            log.warning(f"Companies blacklist has duplicates: {duplicated_names_part}")

        self._data_rows = blacklist_data
        self._names = blacklist_names

    def update(self, rows: Iterable[CompanyBlacklistRow]):
        for row in rows:
            if row.company_name not in self._names:
                table_row = {
                    CompaniesBlacklistColumnNames.COMPANY_NAME.value: row.company_name,
                    CompaniesBlacklistColumnNames.COMPANY_LINK.value: row.company_link,
                    CompaniesBlacklistColumnNames.REASON.value: row.reason
                }
                self._sheet.append(table_row)
                self._data_rows.append(row)
                self._names.add(row.company_name)
            else:
                log.warning(f"Denied vacancy data duplicate ignored: id={row.company_name}")
