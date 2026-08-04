import logging
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .entities import ApplicationDataRow, CompanyBlacklistRow, \
    DeniedVacancyRow, SheetNames
from .errors import FileSchemaError
from .tables import Applications, DeniedVacancies, CompaniesBlacklist

log = logging.getLogger(__name__)


class TableManager:
    """
    Doesn't work with empty files!
    """
    # todo: _create_file_structure_if_empty
    def __init__(self, path: Path):
        self._path = path
        self._work_book: Workbook
        self._open()

    def _open(self):
        self._assert_if_file_blocked()
        self._work_book = load_workbook(self._path)
        self._validate_file_schema()
        self.applications_table = Applications(
            self._work_book[SheetNames.APPLICATIONS.sheet_name]
        )
        self.denied_vacancies_table = DeniedVacancies(
            self._work_book[SheetNames.DENIED_VACANCIES.sheet_name]
        )
        self.companies_blacklist_table = CompaniesBlacklist(
            self._work_book[SheetNames.COMPANIES_BLACKLIST.sheet_name]
        )

    def _assert_if_file_blocked(self):
        with self._path.open("r+"):
            pass

    def _validate_file_schema(self):
        sheet_names = self._work_book.sheetnames
        file_schema = [
            item.sheet_name
            for item in SheetNames
            if item.sheet_name in sheet_names
        ]

        if not len(sheet_names) == len(file_schema):
            names_part = ", ".join(
                f"'{name}'"
                for name in file_schema
                if name not in sheet_names
            )
            raise FileSchemaError(
                f"Invalid file structure. Absent sheet names: {names_part}."
            )

        return file_schema

    def save_data(self):
        temp_path = self._path.with_suffix(f".{datetime.now().strftime('%Y%b%d%H%M%S%f')}.tmp.xlsx")

        self._work_book.save(temp_path)
        try:
            temp_path.replace(self._path)
        except PermissionError as e:
            log.error(
                f"Application data file {self._path} is blocked. "
                f"Current data saved to {temp_path}."
            )
            raise e
        log.info(f"All analyzed data saved successfully to {self._path}.")

    def get_all_ids(self) -> set[str]:
        return self.applications_table.ids | self.denied_vacancies_table.ids

    def contains_id(self, id: str):
        return id in self.applications_table or id in self.denied_vacancies_table.ids

